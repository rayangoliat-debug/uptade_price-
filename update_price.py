import pandas as pd
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import re
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==================== CONFIGURATION ====================
FICHIER_EXCEL = "Fichier comparateur de prix.xlsx"
FEUILLE_HISTORIQUE = "HistoriquePrix"

# URLs exploitables (scraping automatique)
URLS_AUTO = {
    "LeBonCoin Marketplace": "https://www.leboncoin.fr/recherche?text=container+occasion",
    "eBay France": "https://www.ebay.fr/sch/i.html?_nkw=container+occasion",
    "ManoMano": "https://www.manomano.fr/cat/container+stockage",
    "Amazon Business": "https://www.amazon.fr/s?k=container+occasion"
}

# Fournisseurs qui nécessitent une saisie manuelle
FOURNISSEURS_MANUELS = [
    "Nord Container", "2M Containers", "Easy Container", "BBox Container",
    "Méditerranée Containers", "ABC Container", "Est Container", "Ouest Container",
    "IDF Containers", "Toulouse Container"
]

# ==================== SCRAPING ====================
def scraper_leboncoin(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        prix_elements = soup.select('[data-test-id="price"]')
        prix = []
        for el in prix_elements[:3]:
            texte = re.sub(r'[^0-9]', '', el.text)
            if texte:
                prix.append(int(texte))
        return min(prix) if prix else None
    except:
        return None

def scraper_ebay(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        prix_elements = soup.select('.s-item__price')
        prix = []
        for el in prix_elements[:3]:
            texte = re.sub(r'[^0-9]', '', el.text)
            if texte:
                prix.append(int(texte))
        return min(prix) if prix else None
    except:
        return None

def scraper_manomano(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        prix_element = soup.select_one('.product-price')
        if prix_element:
            texte = re.sub(r'[^0-9]', '', prix_element.text)
            return int(texte) if texte else None
        return None
    except:
        return None

def scraper_amazon(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        prix_element = soup.select_one('.a-price-whole')
        if prix_element:
            return int(prix_element.text.replace(',', ''))
        return None
    except:
        return None

# ==================== MISE À JOUR DU FICHIER ====================
def colorer_fournisseurs_manuels():
    """Colorie en jaune les fournisseurs à saisie manuelle"""
    wb = load_workbook(FICHIER_EXCEL)
    ws = wb[FEUILLE_HISTORIQUE]
    
    # Trouver la colonne Fournisseur
    col_fournisseur = None
    for col in range(1, 20):
        if ws.cell(1, col).value == "Fournisseur":
            col_fournisseur = col
            break
    
    if not col_fournisseur:
        print("❌ Colonne 'Fournisseur' non trouvée")
        return
    
    # Couleur jaune clair
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    count = 0
    for row in range(2, ws.max_row + 1):
        fournisseur = ws.cell(row, col_fournisseur).value
        if fournisseur in FOURNISSEURS_MANUELS:
            ws.cell(row, col_fournisseur).fill = yellow_fill
            count += 1
    
    wb.save(FICHIER_EXCEL)
    print(f"✅ {count} fournisseurs colorés en jaune (saisie manuelle)")

def mettre_a_jour_prix_auto():
    """Met à jour les prix des fournisseurs avec scraping auto"""
    
    print("📂 Lecture du fichier...")
    df = pd.read_excel(FICHIER_EXCEL, sheet_name=FEUILLE_HISTORIQUE)
    
    nouveaux_prix = []
    timestamp = datetime.now()
    
    scraper_functions = {
        "LeBonCoin Marketplace": scraper_leboncoin,
        "eBay France": scraper_ebay,
        "ManoMano": scraper_manomano,
        "Amazon Business": scraper_amazon
    }
    
    for fournisseur, scraper_func in scraper_functions.items():
        if fournisseur not in URLS_AUTO:
            continue
        
        url = URLS_AUTO[fournisseur]
        print(f"🔍 Scraping {fournisseur}...")
        prix = scraper_func(url)
        
        if prix:
            for type_container in ["20_occ", "20_neuf", "40_occ", "40_neuf"]:
                # Ajustement approximatif selon le type
                if type_container == "20_neuf":
                    prix_ajuste = int(prix * 1.8)
                elif type_container == "40_occ":
                    prix_ajuste = int(prix * 1.5)
                elif type_container == "40_neuf":
                    prix_ajuste = int(prix * 2.2)
                else:
                    prix_ajuste = prix
                
                nouvelle_ligne = {
                    'Timestamp': timestamp,
                    'Fournisseur': fournisseur,
                    'Région': 'all',
                    'Type Container': type_container,
                    'Prix TTC': prix_ajuste,
                    'Livraison': 0,
                    'Garantie': "selon vendeur",
                    'Délai': "variable",
                    'Source': "marketplace",
                    'Note': 4.0,
                    'Variation %': 0
                }
                nouveaux_prix.append(nouvelle_ligne)
            print(f"   ✅ Prix trouvé: {prix}€ (base)")
        else:
            print(f"   ⚠️ Aucun prix trouvé")
        
        time.sleep(2)
    
    if nouveaux_prix:
        df_nouveaux = pd.DataFrame(nouveaux_prix)
        df_final = pd.concat([df, df_nouveaux], ignore_index=True)
        
        with pd.ExcelWriter(FICHIER_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_final.to_excel(writer, sheet_name=FEUILLE_HISTORIQUE, index=False)
        
        print(f"\n✅ {len(nouveaux_prix)} prix automatiques ajoutés")
    else:
        print("⚠️ Aucun prix automatique ajouté")

def afficher_statut():
    """Affiche les fournisseurs manquants et leur statut"""
    
    df = pd.read_excel(FICHIER_EXCEL, sheet_name=FEUILLE_HISTORIQUE)
    
    print("\n" + "="*50)
    print("📊 STATUT DES FOURNISSEURS")
    print("="*50)
    
    for fournisseur in FOURNISSEURS_MANUELS:
        prix_existants = df[df['Fournisseur'] == fournisseur]['Prix TTC'].tolist()
        if prix_existants:
            dernier_prix = prix_existants[-1]
            print(f"🟡 {fournisseur}: dernier prix = {dernier_prix}€ (à mettre à jour manuellement)")
        else:
            print(f"🔴 {fournisseur}: AUCUN PRIX (saisie urgente)")
    
    print("\n" + "="*50)
    print("✅ Fournisseurs automatiques (OK) :")
    for fournisseur in URLS_AUTO.keys():
        print(f"   🤖 {fournisseur}")

# ==================== LANCER ====================
if __name__ == "__main__":
    print("\n" + "="*50)
    print("📦 GESTION DES PRIX FOURNISSEURS")
    print("="*50)
    print("\n1. Colorer les fournisseurs à saisie manuelle")
    print("2. Scraper les prix automatiques (marketplaces)")
    print("3. Afficher le statut des fournisseurs")
    
    choix = input("\nChoix (1/2/3): ")
    
    if choix == "1":
        colorer_fournisseurs_manuels()
    elif choix == "2":
        mettre_a_jour_prix_auto()
    elif choix == "3":
        afficher_statut()
    else:
        print("Choix invalide")
