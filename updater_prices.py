import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import os

# ==================== CONFIGURATION ====================
FICHIER_EXCEL = "Fichier comparateur de prix.xlsx"
FEUILLE_HISTORIQUE = "HistoriquePrix"

# Sites à scraper avec leurs configurations
SITES_A_SCRAPER = {
    "Box'Innov": {
        "url": "https://www.boxinnov.com/conteneurs-maritimes/conteneur-6-pieds-dry/",
        "type": "unitaire",
        "selecteur_prix": ".price, .product-price",
        "selecteur_nom": "h1, .product-title",
        "region": "all",
        "source": "fournisseur_direct"
    },
    "CFC": {
        "url": "https://compagnie-francaise-du-conteneur.fr/produits/20-pieds-dry",
        "type": "generique",
        "motif_prix": r"(\d{1,3}(?:[\s]?\d{3})?)\s?[€&euro;]\s?TTC",
        "selecteur_nom": "h1, .product-title",
        "region": "all",
        "source": "fournisseur_direct"
    },
    "Eurobox": {
        "url": "https://eurobox.fr/categorie-produit/containers/containers-maritime/",
        "type": "generique",
        "motif_prix": r"(\d{1,3}(?:[\s]?\d{3})?)\s?[€&euro;]\s?HT",
        "selecteur_nom": "h2, h3, .product-title",
        "region": "all",
        "source": "fournisseur_direct"
    },
    "MouvBox": {
        "url": "https://mouvbox-france.com/categorie-produit/containers/les-standards/",
        "type": "generique",
        "motif_prix": r"(?:dès|à partir de)\s*(\d{1,3}(?:[\s]?\d{3})?)\s?[€&euro;]",
        "selecteur_nom": "h2, h3, .product-title",
        "region": "all",
        "source": "fournisseur_direct"
    },
    "Cubner": {
        "url": "https://cubner.com/categorie-produit/conteneur-dry/",
        "type": "generique",
        "motif_prix": r"(?:à partir de|dès)\s*(\d{1,3}(?:[\s]?\d{3})?)\s?[€&euro;]",
        "selecteur_nom": "h2, h3, .product-title",
        "region": "all",
        "source": "fournisseur_direct"
    }
}

# Fournisseurs à saisie manuelle (colorés en jaune)
FOURNISSEURS_MANUELS = [
    "Nord Container", "2M Containers", "Easy Container", "BBox Container",
    "Méditerranée Containers", "ABC Container", "Est Container", "Ouest Container",
    "IDF Containers", "Toulouse Container", "Resotainer", "TITAN Containers France",
    "Bluetainer", "ContainerZ", "Sea Box Company"
]

# ==================== FONCTIONS DE SCRAPING ====================
def get_nom_produit(soup, selecteur):
    """Extrait le nom du produit depuis le HTML"""
    try:
        if selecteur:
            elements = soup.select(selecteur)
            for el in elements:
                nom = el.get_text(strip=True)
                if nom and len(nom) > 3 and not nom.startswith('#'):
                    return nom[:100]  # Limiter la longueur
        # Fallback: chercher dans le titre de la page
        titre = soup.find('title')
        if titre:
            return titre.get_text(strip=True)[:100]
        return "Produit inconnu"
    except:
        return "Produit inconnu"

def scraper_boxinnov(url):
    """Scrape le prix unitaire et le nom sur Box'Innov"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Prix
        texte = soup.get_text()
        match = re.search(r'(\d{1,3}(?:[\s]?\d{3})?)\s?[€&euro;]', texte)
        prix = int(match.group(1).replace(' ', '')) if match else None
        
        # Nom du produit
        nom = get_nom_produit(soup, "h1, .product-title, .title")
        
        return prix, nom
    except Exception as e:
        print(f"   Erreur Box'Innov: {e}")
        return None, None

def scraper_site_generique(config):
    """Scrape le prix générique et les noms de produits"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(config["url"], headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Prix
        texte = soup.get_text()
        match = re.search(config["motif_prix"], texte, re.I)
        prix = int(match.group(1).replace(' ', '')) if match else None
        
        # Noms des produits
        noms_produits = []
        selecteurs = ["h2", "h3", ".product-title", ".product-name", ".title"]
        
        for sel in selecteurs:
            elements = soup.select(sel)
            for el in elements:
                nom = el.get_text(strip=True)
                if nom and len(nom) > 5 and len(nom) < 100:
                    if any(mot in nom.lower() for mot in ['container', 'conteneur', 'pieds', 'dry', '20', '40']):
                        noms_produits.append(nom)
        
        # Supprimer les doublons
        noms_produits = list(dict.fromkeys(noms_produits))
        
        if not noms_produits:
            noms_produits = ["Container standard"]
        
        return prix, noms_produits
    except Exception as e:
        print(f"   Erreur: {e}")
        return None, []

# ==================== MISE À JOUR EXCEL ====================
def colorer_fournisseurs_manuels():
    """Colore en jaune les fournisseurs à saisie manuelle"""
    try:
        wb = openpyxl.load_workbook(FICHIER_EXCEL)
        ws = wb[FEUILLE_HISTORIQUE]
        
        col_fournisseur = None
        for col in range(1, 20):
            if ws.cell(1, col).value == "Fournisseur":
                col_fournisseur = col
                break
        
        if not col_fournisseur:
            print("❌ Colonne 'Fournisseur' non trouvée")
            return
        
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        count = 0
        
        for row in range(2, ws.max_row + 1):
            fournisseur = ws.cell(row, col_fournisseur).value
            if fournisseur in FOURNISSEURS_MANUELS:
                ws.cell(row, col_fournisseur).fill = yellow_fill
                count += 1
        
        wb.save(FICHIER_EXCEL)
        print(f"✅ {count} fournisseurs colorés en jaune")
    except Exception as e:
        print(f"Erreur coloration: {e}")

def mettre_a_jour_prix():
    """Scrape tous les sites et met à jour le fichier Excel"""
    
    print("\n📂 Mise à jour des prix...")
    
    try:
        df = pd.read_excel(FICHIER_EXCEL, sheet_name=FEUILLE_HISTORIQUE)
    except:
        df = pd.DataFrame(columns=['Timestamp', 'Fournisseur', 'Produit', 'Région', 'Type Container', 
                                   'Prix TTC', 'Livraison', 'Garantie', 'Délai', 'Source', 'Note'])
    
    nouvelles_lignes = []
    timestamp = datetime.now()
    
    for fournisseur, config in SITES_A_SCRAPER.items():
        print(f"\n🔍 Scraping {fournisseur}...")
        
        if config["type"] == "unitaire":
            prix, nom = scraper_boxinnov(config["url"])
            noms_produits = [nom] if nom else ["Container"]
        else:
            prix, noms_produits = scraper_site_generique(config)
        
        if prix and noms_produits:
            for nom_produit in noms_produits[:5]:  # Limiter à 5 produits par site
                # Déterminer le type de container à partir du nom
                type_container = "20_occ"
                if "20" in nom_produit and "neuf" in nom_produit.lower():
                    type_container = "20_neuf"
                elif "40" in nom_produit and "neuf" in nom_produit.lower():
                    type_container = "40_neuf"
                elif "40" in nom_produit:
                    type_container = "40_occ"
                elif "20" in nom_produit:
                    type_container = "20_occ"
                
                nouvelle_ligne = {
                    'Timestamp': timestamp,
                    'Fournisseur': fournisseur,
                    'Produit': nom_produit,
                    'Région': config["region"],
                    'Type Container': type_container,
                    'Prix TTC': prix,
                    'Livraison': 0,
                    'Garantie': "selon fournisseur",
                    'Délai': "variable",
                    'Source': config["source"],
                    'Note': 4.0
                }
                nouvelles_lignes.append(nouvelle_ligne)
            print(f"   ✅ {len(noms_produits)} produits trouvés - Prix: {prix} €")
        else:
            print(f"   ⚠️ Aucun prix trouvé")
    
    # Ajouter les nouvelles lignes
    if nouvelles_lignes:
        df_nouvelles = pd.DataFrame(nouvelles_lignes)
        df_final = pd.concat([df, df_nouvelles], ignore_index=True)
        
        with pd.ExcelWriter(FICHIER_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_final.to_excel(writer, sheet_name=FEUILLE_HISTORIQUE, index=False)
        
        print(f"\n✅ {len(nouvelles_lignes)} lignes ajoutées")
        colorer_fournisseurs_manuels()
    else:
        print("⚠️ Aucune donnée ajoutée")

# ==================== LANCER ====================
if __name__ == "__main__":
    print("\n" + "="*50)
    print("📦 SCRAPING DES PRIX CONTAINERS")
    print("="*50)
    
    if not os.path.exists(FICHIER_EXCEL):
        print(f"❌ Fichier {FICHIER_EXCEL} non trouvé")
        exit()
    
    mettre_a_jour_prix()
    print("\n✅ Script terminé")
